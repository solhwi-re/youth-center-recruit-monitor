from openpyxl import load_workbook
from .config import DATA_PATH
from .models import Center


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_centers() -> list[Center]:
    wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
    ws = wb.active
    centers: list[Center] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        # A:순번 B:광역시도 C:시군구 D:공간명 E:도로명주소 F:상세주소 G:URL H:운영법인
        region1 = clean(row[1] if len(row) > 1 else "")
        region2 = clean(row[2] if len(row) > 2 else "")
        center_name = clean(row[3] if len(row) > 3 else "")
        address = " ".join([clean(row[4] if len(row) > 4 else ""), clean(row[5] if len(row) > 5 else "")]).strip()
        homepage_url = clean(row[6] if len(row) > 6 else "")
        operator_name = clean(row[7] if len(row) > 7 else "")
        if not center_name:
            continue
        centers.append(Center(
            center_id=idx,
            region1=region1,
            region2=region2,
            center_name=center_name,
            address=address,
            homepage_url=homepage_url,
            operator_name=operator_name,
        ))
    return centers
